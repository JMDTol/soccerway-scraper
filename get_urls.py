from tkinter import filedialog
from selenium import webdriver
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import time
import tkinter as tk


root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()

url = input('Input season URL: ')

driver = webdriver.Chrome()
driver.fullscreen_window()
driver.get(url)

time.sleep(0.5)

if driver.find_element_by_class_name('qc-cmp-button'):
    privacy = driver.find_element_by_class_name('qc-cmp-button')
    privacy.click()

time.sleep(0.5)

week = driver.find_element_by_id('page_competition_1_block_competition_matches_summary_5_1_2')
week.click()

time.sleep(0.5)

html = driver.find_element_by_tag_name('html').get_attribute('innerHTML')
game_week_soup = BeautifulSoup(html, 'html.parser')

previous_clicks = 0
url_list = []

for week in game_week_soup.findAll(id='page_competition_1_block_competition_matches_summary_5_page_dropdown'):
    number_weeks = (week.contents[-1])
    previous_clicks = int(number_weeks.contents[0]) - 1

for info in game_week_soup.findAll('td', class_='info-button button'):
    for link in info.find_all('a', href=True):
        match_url = 'https://us.soccerway.com' + link.get('href')
        url_list.append(match_url)

for i in range(previous_clicks):
    previous_button = driver.find_element_by_class_name('previous')
    previous_button.click()
    time.sleep(0.1)
    html = driver.find_element_by_tag_name('html').get_attribute('innerHTML')
    game_week_soup = BeautifulSoup(html, 'html.parser')
    for info in game_week_soup.findAll('td', class_='info-button button'):
        for link in info.find_all('a', href=True):
            match_url = 'https://us.soccerway.com' + link.get('href')
            url_list.append(match_url)

driver.quit()

print('=' * 100 + '\n{} matches added to {}'.format(len(url_list), file_path))

wb = load_workbook(file_path)
ws = wb.active

column_length = 0

for i in ws.rows:
    column_length += 1

url_list_reversed = url_list[::-1]

for url in url_list_reversed:
    ws.cell(row=column_length, column=1).value = url
    column_length += 1

wb.save(file_path)
